import csv
import openpyxl
from openpyxl import Workbook



wb = Workbook()
wb.create_sheet('name')
wb['name']['A1'].value, wb['name']['B1'].value = 1, 'One'
wb['name']['A2'].value, wb['name']['B2'].value = 2, 'Two'
wb['name']['A3'].value, wb['name']['B3'].value = 3, 'Three'
wb['name']['A4'].value, wb['name']['B4'].value = 4, 'Four'

wb.save('ex.xlsx')

wb = openpyxl.load_workbook('ex.xlsx')
d = []
for i in range(1,5):
    d.append({
        f'A{i}': str(wb['name'][f'A{i}'].value), f'B{i}': str(wb['name'][f'B{i}'].value)
    })
print('Table in list format:', d)


Headers = []
for i in range(1, 5):
    Headers.append(f'A{i}')
    Headers.append(f'B{i}')
print('Headers:', Headers)
with open('ex.csv', 'w') as f:
    writer = csv.DictWriter(f, fieldnames=Headers)
    writer.writerows(d)



